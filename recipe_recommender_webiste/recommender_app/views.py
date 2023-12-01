from django.shortcuts import render, redirect
import pickle
import pandas as pd
from recommender_app.recommendation import *
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import UserCreationForm
from .models import Liked_Recipes, Recipe_Data
import json 
from openpyxl import load_workbook
from datetime import datetime

num1,num2 = 0,10

def signup(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=raw_password)
            login(request, user)
            return redirect('/recommender_app/new_user/')
    else:
        form = UserCreationForm()
    return render(request, 'signup.html', {'form': form})


def first_time_dashboard(request):
    user = request.user.username

    return render(request, 'new_user_dash.html', {"user":user})


def recommend(request):

    with open(r'D:\recipe_recommender\recipe_recommender_webiste\recommender_app\static\ingredients2.pickle','rb') as f:
        ingreds = pickle.load(f)

    ingreds2 = [str(i).title() for i in ingreds]

    d = {}
    for i in range(len(ingreds)):
        d[ingreds[i]] = ingreds2[i] 


    ethnicity = ['spanish','egyptian','tex-mex','welsh','australian','polynesian','pacific-northwest',
    'swiss','italian','german','sudanese','chilean','costa-rican','hawaiian','malaysian','icelandic','norwegian',
    'pennsylvania-dutch','somalian','pakistani','nepalese','venezuelan','asian','puerto-rican','nigerian',
    'ethiopian','vietnamese','hungarian','laotian','dutch','chinese','south-african','brazilian','austrian',
    'greek','middle-eastern','iraqi','ecuadorean','creole','african','korean','cambodian','caribbean','thai',
    'indonesian','south-american','midwestern','north-american','mongolian','libyan','new-zealand','russian',
    'congolese','mexican','south-west-pacific','japanese','native-american','portuguese','palestinian','cuban',
    'scottish','central-american','filipino','indian','honduran','european','saudi-arabian','guatemalan','argentine',
    'irish','turkish','georgian','british-columbian','peruvian','scandinavian','southern-united-states','lebanese',
    'czech','canadian','cajun','northeastern-united-states','colombian','southwestern-united-states']

    ethnicity2 = [i.title() for i in ethnicity] 
    eth_d = {} 
    for i in range(len(ethnicity)): 
        eth_d[ethnicity[i]] = ethnicity2[i] 

    types_of_food = ['spreads','omelets-and-frittatas','frozen-desserts','dips','salsas',
    'shakes','appetizers','sandwiches','cobblers-and-crisps','soul',
    'seafood','pies','salads','barbecue','stews','side-dishes']

    types_of_food2 = [i.title() for i in types_of_food]
    types_d = {}
    for i in range(len(types_of_food)):
        types_d[types_of_food[i]] = types_of_food2[i]

    holiday = ['super-bowl','rosh-hashanah','new-years','mothers-day','st-patricks-day',
    'fathers-day','valentines-day','ramadan','chinese-new-year','easter','christmas',
    'april-fools-day','cinco-de-mayo','independence-day','labor-day','memorial-day',
    'passover','kwanzaa','thanksgiving','hanukkah','halloween']

    holiday2 = [i.title() for i in holiday]

    hol_d = {}
    for i in range(len(holiday)):
        hol_d[holiday[i]] = holiday2[i]

    dietary_restrict = ['vegan','vegetarian','dairy-free','gluten-free','nut-free']
    dietary_restrict2 = [i.title() for i in dietary_restrict]

    dietary_restrict_d = {}
    for i in range(len(dietary_restrict)):
        dietary_restrict_d[dietary_restrict[i]] = dietary_restrict2[i]


    return render(request,'recommend.html',{'ingreds':d,'ethnicity':eth_d,'food_type':types_d,'holiday':hol_d,
        'dietary':dietary_restrict_d })



def results(request):
    global num1
    global num2

    like_button = request.POST.get("like_button")
    show_more = request.POST.get("show_more")
    
    calories = request.POST.get('calories')
    sugar = request.POST.get("sugar")
    protein = request.POST.get("protein")
    carbs = request.POST.get("carbs")
    time = request.POST.get("time")

    if (show_more == "yes") or (like_button == "yes"):

        dietary = request.POST.get("dietary")
        eth = request.POST.get("ethnicity")
        ingredients = request.POST.get('ingredients')

        dietary = ast.literal_eval(dietary)
        eth = ast.literal_eval(eth)
        ingredients = ast.literal_eval(ingredients)

        if show_more == "yes":
            num1 += 10
            num2 += 10

        if like_button == "yes":
            recipe_row = request.POST.get("recipe_row")
            recipe_name_row = request.POST.get("recipe_name_row")
            ingredients_row = request.POST.get("ingredients_row")
            description_row = request.POST.get("description_row")
            steps_row = request.POST.get("steps_row")
            calories_row = request.POST.get("calories_row")
            protein_row = request.POST.get("protein_row")
            carbs_row = request.POST.get("carbs_row")
            sugar_row = request.POST.get("sugar_row")
            time_to_cook_row = request.POST.get("time_to_cook_row")
            tag_bins_row = request.POST.get("tag_bins_row")
            ingred_bins_row = request.POST.get("ingred_bins_row")
            user = request.user.username
            date_liked = datetime.today()
            
            liked_recipe = Liked_Recipes(recipe=recipe_row,user=user,
                    recipe_name=recipe_name_row,description=description_row,ingredients=ingredients_row,
                    steps=steps_row,calories=calories_row,protein=protein_row,carbs=carbs_row,sugar=sugar_row,
                    time=time_to_cook_row,tag_bins=tag_bins_row,ingredients_bin=ingred_bins_row,date_liked=date_liked)
            
            liked_recipe.save()

            #print(Liked_Recipes.objects.all())
            

    ## Use pagination instead

    elif (show_more != "yes") and (like_button != "yes"):
        num1,num2 = 0,10

        ingredients = request.POST.getlist('ingredients')
        dietary = request.POST.getlist("dietary")
        eth = request.POST.getlist("ethnicity")

        
        
    d = {'ingredients_selected': [ingredients], 'calories': [calories],'protein':[protein],'carbs':[carbs],'sugar':[sugar],'time_filter':[time],'dietary_restriction':[dietary],
        'ethnicity':[eth],'dietary':[dietary],'time':[time]}

    df = filter_df(d)
    out_df = prepare_df(df,num1,num2)
    data = out_df.all()
    
    return render(request,'results.html',{'recs': data,'ingredients':ingredients,'calories':calories,'sugar':sugar,
                                          'protein':protein,'time':time,'dietary':dietary,'ethnicity':eth})



def dashboard(request):
    new_like = request.POST.get("new_like")

    if new_like:
        new_like = int(new_like)
        recipe = Recipe_Data.objects.filter(recipe=new_like).values_list('recipe',flat=True)[0]
        recipe_name = Recipe_Data.objects.filter(recipe=new_like).values_list('recipe_name_html',flat=True)[0]
        description = Recipe_Data.objects.filter(recipe=new_like).values_list('description_html',flat=True)[0]
        ingredients = Recipe_Data.objects.filter(recipe=new_like).values_list('ingredients_html',flat=True)[0]
        steps = Recipe_Data.objects.filter(recipe=new_like).values_list('steps_html',flat=True)[0]
        calories = Recipe_Data.objects.filter(recipe=new_like).values_list('calories',flat=True)[0]
        protein = Recipe_Data.objects.filter(recipe=new_like).values_list('protein',flat=True)[0]
        carbs = Recipe_Data.objects.filter(recipe=new_like).values_list('carbs',flat=True)[0]
        sugar = Recipe_Data.objects.filter(recipe=new_like).values_list('sugar',flat=True)[0]
        time = Recipe_Data.objects.filter(recipe=new_like).values_list('minutes',flat=True)[0]
        tag_bins = Recipe_Data.objects.filter(recipe=new_like).values_list('tag_bins',flat=True)[0]
        ingredients_bin = Recipe_Data.objects.filter(recipe=new_like).values_list('ingredients_bin',flat=True)[0]
        user=request.user.username
        date_liked = datetime.today()

        print(recipe)
        print(recipe_name)

        liked_recipe = Liked_Recipes(recipe=recipe,user=user,
                    recipe_name=recipe_name,description=description,ingredients=ingredients,
                    steps=steps,calories=calories,protein=protein,carbs=carbs,sugar=sugar,
                    time=time,tag_bins=tag_bins,ingredients_bin=ingredients_bin,date_liked=date_liked)
            
        liked_recipe.save()


    recipes = Liked_Recipes.objects.filter(user=request.user.username).order_by('-date_liked').all()[0:3]
    user = request.user.username
    
    recs = getNeighbors(recipes)
    out_rec = Recipe_Data.objects.filter(recipe__in=recs).all()

    return render(request,'dashboard.html',{"recipes":recipes,"user":user,"recs":out_rec})


def liked_recipes(request):
    liked = Liked_Recipes.objects.filter(user=request.user.username).order_by('-date_liked').all()

    return render(request,'liked_recipes.html',{"liked":liked})



def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            recipe, tags, nutrition, minutes,ingredients,ethnicity,type_of_food,holiday,dietary_restriction,ingredients_bin,calories,sugar,protein,carbs,recipe_name,recipe_steps,description,ingredients_html,steps_html,description_html,recipe_name_html,tag_bins= row
            Recipe_Data.objects.create(recipe=recipe, tags=tags, nutrition=nutrition,minutes=minutes,ingredients=ingredients,
                    ethnicity=ethnicity,type_of_food=type_of_food,holiday=holiday,dietary_restriction=dietary_restriction,
                    ingredients_bin=ingredients_bin,calories=calories,sugar=sugar,protein=protein,carbs=carbs,
                    recipe_name=recipe_name,recipe_steps=recipe_steps,description=description,
                    ingredients_html=ingredients_html,steps_html=steps_html,description_html=description_html,recipe_name_html=recipe_name_html,tag_bins=tag_bins)

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')





    
